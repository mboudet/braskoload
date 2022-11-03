import pandas as pd
import os
from openpyxl import load_workbook
from checkcel import Checkerator

envoi_file = "/home/mboudet/Téléchargements/info_envoi.xlsx"
os.makedirs("outputs", exist_ok=True)


def generate_file_name(file, entity, organism):
    entity = entity.replace(", ", "_").lower()
    file_name = os.path.basename(file)
    name, ext = os.path.splitext(file_name)
    if not organism or organism in name:
        return "{}_{}{}".format(name, entity, ext)

    return "{}_{}_{}{}".format(name, organism, entity, ext)


def generate_id(pop_name, input_file, entity, block=None, plant=None):
    entity = entity.replace(", ", "_").lower()
    file_suffix = input_file.split("_")[1]
    base_id = "{}_{}_{}".format(pop_name, file_suffix, entity)

    if block:
        base_id = "{}_{}".format(base_id, block)
    if plant:
        base_id = "{}_{}".format(base_id, plant)

    return base_id


def write_row(ws, row, input_file, entity, population, block, plant):
    id = generate_id(population, input_file, entity, block, plant)
    ws["A{}".format(row)] = id
    ws["B{}".format(row)] = population
    ws["C{}".format(row)] = entity

    if block:
        ws["D{}".format(row)] = block
        if plant:
            ws["E{}".format(row)] = plant


input_data = {
    "Brassica rapa": {
        "sheet": 1,
        "suffix": "BR",
        "input_files": [
            "BrasExplor_emergence_sheet_BR.xlsx",
            "BrasExplor_flowering_sheet.xlsx",
            "BrasExplor_harvest_sheet.xlsx",
            "BrasExplor_leaves_sheet_BR.xlsx",
        ]
    },
    "Brassica oleracea": {
        "sheet": 0,
        "suffix": "BO",
        "input_files": [
            "BrasExplor_emergence_sheet_BO.xlsx",
            "BrasExplor_flowering_sheet.xlsx",
            "BrasExplor_harvest_sheet.xlsx",
            "BrasExplor_leaves_sheet_BO.xlsx",
        ]
    }
}

base_template = "/home/mboudet/Bureau/checkcel_templates/BrasExplor/WP3"

for val in input_data.values():
    for output in val['input_files']:
        check = Checkerator(output=output)
        check.load_from_python_file(os.path.join(base_template, output.replace("xlsx", "py")))
        check.generate()

# Also generate soil
check = Checkerator(output="BrasExplor_Soil_analysis_sheet.xlsx")
check.load_from_python_file(os.path.join(base_template, "BrasExplor_Soil_analysis_sheet.py"))
check.generate()

sites = set()

for key, value in input_data.items():
    org = value["suffix"]
    pop_dict = {}
    df = pd.read_excel(envoi_file, value["sheet"])
    for col in df.columns:
        pop_dict[col] = [val for val in df[col].dropna() if val]
        sites.add(col)

    for input_file in value['input_files']:
        for entity in pop_dict:
            row = 2

            new_file_name = generate_file_name(input_file, entity, org)
            os.makedirs(os.path.join("outputs", entity), exist_ok=True)

            wb = load_workbook(input_file)
            ws = wb['Data']

            has_block = ws['D1'].value == "Block"
            has_plant = ws['E1'].value == "Plant number"

            for pop in pop_dict[col]:
                if has_block:
                    for block in range(1, 4):
                        if has_plant:
                            for plant in range(1, 6):
                                write_row(ws, row, input_file, entity, pop, block, plant)
                                row += 1
                        else:
                            write_row(ws, row, input_file, entity, pop, block, None)
                            row += 1
                else:
                    write_row(ws, row, input_file, entity, pop, None, None)
                    row += 1

            wb.save(filename=os.path.join("outputs", entity, new_file_name))

for col in sites:
    new_file_name = generate_file_name("BrasExplor_Soil_analysis_sheet.xlsx", col, None)
    wb = load_workbook("BrasExplor_Soil_analysis_sheet.xlsx")
    ws = wb['Data']
    ws['A2'] = col
    wb.save(os.path.join("outputs", col, new_file_name))
